#!/usr/bin/env python3
"""Build workbook output plan items from Linux/Unix SSH command evidence."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.utils.cell import range_boundaries
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl is required to build SSH workbook plans") from exc

from analyze_checklist import (
    ADMIN_INTERVIEW_FINDING,
    ADMIN_INTERVIEW_RESULT,
    find_header,
    is_admin_interview,
    norm,
    row_text,
)


DEFAULT_FINDING_TEMPLATE = "{name}"
DEFAULT_SCREENSHOT_FINDING_TEMPLATE = "{name}"
DEFAULT_RESULT_OPTIONS = ("不符合", "符合", "不适用", "未检查")
DEFAULT_RESULT_TEXT = ""


def safe_file_part(text: str) -> str:
    if not text.strip():
        return "command"
    value = re.sub(r'[\\/:*?"<>|\s]+', "_", text)
    value = re.sub(r"[^A-Za-z0-9_.-]", "_", value)
    return value.strip("_") or "command"


def load_json_array(path: Path, label: str) -> list[dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    if isinstance(data, dict):
        data = data.get("items") or data.get("commands") or data.get("results") or [data]
    if not isinstance(data, list):
        raise SystemExit(f"{label} must be a JSON array or an object containing an array")
    return [dict(item) for item in data]


def command_row(command: dict[str, Any]) -> int | None:
    for key in ("row", "workbookRow", "workbook_row"):
        if command.get(key) not in (None, ""):
            return int(command[key])
    combined = " ".join(str(command.get(key) or "") for key in ("id", "name", "command"))
    match = re.search(r"\brow\s*0*(\d{1,4})\b", combined, flags=re.I)
    if match:
        return int(match.group(1))
    return None


def manifest_by_id(manifest_path: Path | None) -> dict[str, dict[str, Any]]:
    if not manifest_path:
        return {}
    entries = load_json_array(manifest_path, "manifest")
    by_id: dict[str, dict[str, Any]] = {}
    for entry in entries:
        entry_id = safe_file_part(str(entry.get("id") or ""))
        if entry_id:
            by_id[entry_id] = entry
    return by_id


def evidence_name(index: int, command: dict[str, Any], manifest: dict[str, dict[str, Any]]) -> str:
    command_id = safe_file_part(str(command.get("id") or f"cmd{index:02d}"))
    manifest_entry = manifest.get(command_id)
    if manifest_entry and manifest_entry.get("screenshot"):
        return Path(str(manifest_entry["screenshot"])).name
    if command.get("evidence"):
        evidence = command["evidence"]
        if isinstance(evidence, list) and evidence:
            return Path(str(evidence[0])).name
        return Path(str(evidence)).name
    return f"{index:02d}_{command_id}.png"


def command_artifact(command_id: str, command: dict[str, Any], manifest: dict[str, dict[str, Any]]) -> str:
    manifest_entry = manifest.get(command_id) or {}
    for key in ("output", "stdout", "transcript"):
        if manifest_entry.get(key):
            return Path(str(manifest_entry[key])).name
    if command.get("output"):
        return Path(str(command["output"])).name
    return ""


def command_artifact_path(
    command_id: str,
    command: dict[str, Any],
    manifest: dict[str, dict[str, Any]],
    manifest_json: Path | None,
) -> Path | None:
    manifest_entry = manifest.get(command_id) or {}
    value = None
    for key in ("output", "stdout", "transcript"):
        if manifest_entry.get(key):
            value = str(manifest_entry[key])
            break
    if value is None and command.get("output"):
        value = str(command["output"])
    if not value:
        return None
    path = Path(value)
    if not path.is_absolute() and manifest_json:
        path = manifest_json.parent / path
    return path


def command_output_text(
    command_id: str,
    command: dict[str, Any],
    manifest: dict[str, dict[str, Any]],
    manifest_json: Path | None,
) -> str:
    path = command_artifact_path(command_id, command, manifest, manifest_json)
    if not path or not path.exists():
        return ""
    return path.read_text(encoding="utf-8", errors="replace")


def range_values(sheet: Any, reference: str) -> list[str]:
    workbook = sheet.parent
    reference = reference.strip()
    if reference.startswith("="):
        reference = reference[1:].strip()
    target_sheet = sheet
    coordinate = reference
    if "!" in reference:
        sheet_name, coordinate = reference.split("!", 1)
        sheet_name = sheet_name.strip("'")
        if sheet_name not in workbook.sheetnames:
            return []
        target_sheet = workbook[sheet_name]
    coordinate = coordinate.replace("$", "")
    try:
        min_col, min_row, max_col, max_row = range_boundaries(coordinate)
    except ValueError:
        return []
    values: list[str] = []
    for row in target_sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            value = norm(cell.value)
            if value:
                values.append(value)
    return values


def named_range_values(sheet: Any, name: str) -> list[str]:
    workbook = sheet.parent
    name = name.lstrip("=").strip()
    defined_names: list[Any] = []
    try:
        defined = workbook.defined_names.get(name)
        if defined:
            defined_names.append(defined)
    except AttributeError:
        pass
    try:
        defined_names.extend(entry for entry in workbook.defined_names.definedName if entry.name == name)
    except AttributeError:
        pass
    values: list[str] = []
    for defined in defined_names:
        try:
            destinations = list(defined.destinations)
        except Exception:
            continue
        for sheet_name, coordinate in destinations:
            if sheet_name in workbook.sheetnames:
                values.extend(range_values(workbook[sheet_name], coordinate))
    return values


def validation_options(sheet: Any, formula: str) -> list[str]:
    formula = formula.strip()
    if formula.startswith('"') and formula.endswith('"'):
        return [part.strip() for part in formula[1:-1].split(",") if part.strip()]
    values = range_values(sheet, formula)
    if values:
        return values
    values = named_range_values(sheet, formula)
    if values:
        return values
    if "," in formula and "!" not in formula and ":" not in formula:
        return [part.strip().strip('"') for part in formula.lstrip("=").split(",") if part.strip()]
    return []


def result_options_for_sheet(sheet: Any, result_col: int | None) -> list[str]:
    if not result_col or not getattr(sheet, "data_validations", None):
        return list(DEFAULT_RESULT_OPTIONS)
    options: list[str] = []
    for validation in sheet.data_validations.dataValidation:
        if validation.type != "list" or not validation.formula1:
            continue
        applies_to_result_col = any(
            cell_range.min_col <= result_col <= cell_range.max_col
            for cell_range in validation.cells.ranges
        )
        if not applies_to_result_col:
            continue
        formula = str(validation.formula1).strip()
        options = validation_options(sheet, formula)
        if options:
            break
    return options or list(DEFAULT_RESULT_OPTIONS)


def unchecked_result(options: list[str]) -> str:
    if "未检查" in options:
        return "未检查"
    return ""


def normalize_result(value: str, options: list[str], *, strict: bool = False, label: str = "result") -> str:
    text = norm(value)
    if not options:
        return text
    if text in options:
        return text
    mapping = {
        "已检查": "符合",
        "通过": "符合",
        "合规": "符合",
        "符合要求": "符合",
        "不通过": "不符合",
        "不合规": "不符合",
        "不符合要求": "不符合",
        "na": "不适用",
        "n/a": "不适用",
        "不涉及": "不适用",
        "未执行": "未检查",
        "未操作": "未检查",
    }
    mapped = mapping.get(text.lower(), mapping.get(text))
    if mapped and mapped in options:
        return mapped
    if not strict:
        for option in options:
            if option in text or text in option:
                return option
    if strict:
        raise SystemExit(f"{label} {value!r} is not one of the workbook result options: {', '.join(options)}")
    return unchecked_result(options)


def after_marker_lines(text: str, marker: str) -> list[str]:
    if marker not in text:
        return []
    tail = text.split(marker, 1)[1]
    lines = []
    for line in tail.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        if stripped.startswith("END ROW") or stripped.startswith("--- "):
            break
        lines.append(stripped)
    return lines


def clean_lines(text: str) -> list[str]:
    ignored_prefixes = ("$ command", "COMMAND:", "ROW ", "END ROW", "TERM environment")
    lines: list[str] = []
    for line in text.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        if any(stripped.startswith(prefix) for prefix in ignored_prefixes):
            continue
        lines.append(stripped)
    return lines


def pass_values(text: str) -> dict[str, str]:
    values: dict[str, str] = {}
    for name in ("PASS_MAX_DAYS", "PASS_MIN_DAYS", "PASS_MIN_LEN", "PASS_WARN_AGE"):
        match = re.search(rf"^\s*{name}\s+(\S+)", text, flags=re.M)
        if match:
            values[name] = match.group(1)
    return values


def int_value(values: dict[str, str], key: str) -> int | None:
    try:
        return int(values[key])
    except (KeyError, TypeError, ValueError):
        return None


def summarize_linux_output(row: int, output_text: str, options: list[str]) -> tuple[str, str]:
    text = output_text.replace("\r\n", "\n")
    result = "符合"
    finding = ""

    if row == 5:
        statuses = re.findall(r"^([^:\s]+):(EMPTY|LOCKED|SET)$", text, flags=re.M)
        set_users = [user for user, state in statuses if state == "SET"]
        empty_users = [user for user, state in statuses if state == "EMPTY"]
        for user in after_marker_lines(text, "--- empty password accounts ---"):
            if user not in empty_users:
                empty_users.append(user)
        result = "不符合" if empty_users else "符合"
        set_text = "、".join(set_users[:6]) if set_users else "未发现"
        empty_text = "、".join(empty_users) if empty_users else "空"
        finding = f"/etc/shadow 状态显示 {set_text} 为 SET；空口令账户为{empty_text}。"
    elif row == 6:
        values = pass_values(text)
        max_days = int_value(values, "PASS_MAX_DAYS")
        min_len = int_value(values, "PASS_MIN_LEN")
        has_quality = "pam_pwquality" in text or "pam_cracklib" in text
        result = "符合" if (
            max_days is not None and max_days <= 90
            and min_len is not None and min_len >= 10
            and has_quality
        ) else "不符合"
        parts = [f"{key}={values[key]}" for key in ("PASS_MAX_DAYS", "PASS_MIN_DAYS", "PASS_MIN_LEN", "PASS_WARN_AGE") if key in values]
        quality = "已引用 pam_pwquality" if has_quality else "未发现 PAM 复杂度模块"
        finding = f"/etc/login.defs 显示 {', '.join(parts) or '未读取到 PASS_* 参数'}；{quality}。"
    elif row == 10:
        has_lock = bool(re.search(r"pam_(faillock|tally2)|deny\s*=", text, flags=re.I))
        result = "符合" if has_lock else "不符合"
        finding = "PAM 已配置登录失败锁定参数。" if has_lock else "未发现 pam_faillock/pam_tally2、deny、unlock_time 等登录失败锁定配置。"
    elif row == 11:
        cleaned = "\n".join(clean_lines(text))
        ssh_active = "active" in cleaned and re.search(r"[:*]22\b", cleaned)
        telnet_open = bool(re.search(r"[:*]23\b|telnet-server", cleaned, flags=re.I))
        result = "符合" if ssh_active and not telnet_open else "不符合"
        finding = "sshd 为 active/enabled，22 端口监听；未发现 telnet 23 端口监听。" if result == "符合" else "远程管理协议检查发现 SSH 未正常启用或存在 telnet 相关痕迹。"
    elif row == 12:
        lines = clean_lines(text)
        duplicate_names = []
        duplicate_uids = []
        current = "names"
        for line in lines:
            if line.startswith("--- duplicate UIDs"):
                current = "uids"
                continue
            if line.startswith("--- passwd users"):
                current = "users"
                continue
            if current == "names" and not line.startswith("---"):
                duplicate_names.append(line)
            elif current == "uids" and not line.startswith("---"):
                duplicate_uids.append(line)
        result = "不符合" if duplicate_names or duplicate_uids else "符合"
        finding = "未发现重复用户名或重复 UID。" if result == "符合" else f"发现重复用户名/UID：{', '.join((duplicate_names + duplicate_uids)[:8])}。"
    elif row == 14:
        passwd_match = re.search(r"(?m)^\s*([0-7]{1,4})\s+\S+\s+\S+\s+/etc/passwd\b", text)
        shadow_match = re.search(r"(?m)^\s*([0-7]{1,4})\s+\S+\s+\S+\s+/etc/shadow\b", text)
        passwd_mode = int(passwd_match.group(1), 8) if passwd_match else None
        shadow_mode = int(shadow_match.group(1), 8) if shadow_match else None
        result = "符合" if passwd_mode == 0o644 and shadow_mode is not None and (shadow_mode & 0o077) == 0 else "不符合"
        passwd_display = f"{passwd_mode:03o}" if passwd_mode is not None else "未识别"
        shadow_display = f"{shadow_mode:03o}" if shadow_mode is not None else "未识别"
        finding = f"/etc/passwd 权限为 {passwd_display}；/etc/shadow 权限为 {shadow_display}。"
    elif row == 17:
        root_password = re.search(r"^root\s+PS\b", text, flags=re.M)
        locked_defaults = re.findall(r"^(\w+)\s+LK\b", text, flags=re.M)
        result = "不符合" if root_password else "符合"
        finding = "root 默认账户仍设置口令；nobody 等默认账户为锁定状态。" if root_password else f"默认账户口令状态已检查，锁定账户包括：{', '.join(locked_defaults) or '未识别'}。"
    elif row == 18:
        interactive = []
        for line in clean_lines(text):
            if re.match(r"^[^:]+:\d+:[^:]+:/", line):
                interactive.append(line.split(":", 1)[0])
        suspicious = [user for user in interactive if user not in {"root", "test"}]
        result = "不符合" if suspicious else "符合"
        finding = f"交互式 Shell 账户包括：{', '.join(interactive) or '未发现'}。" + (f" 需确认 {', '.join(suspicious)} 是否为必要账户。" if suspicious else "")
    elif row == 19:
        active = "auditd.service" in text and "active (running)" in text
        result = "符合" if active else "不符合"
        finding = "auditd.service 为 enabled 且 active(running)。" if active else "未确认 auditd 处于 enabled/active(running) 状态。"
    elif row == 20:
        no_rules = "No rules" in text
        auditd_running = "/sbin/auditd" in text or "enabled 1" in text
        result = "不符合" if no_rules else ("符合" if auditd_running else "不符合")
        finding = "auditd 进程存在，auditctl 状态 enabled=1；当前审计规则为空。" if no_rules else "auditd 进程与审计规则已输出。"
    elif row == 21:
        rules = [line for line in clean_lines(text) if not line.startswith("--- audit.rules") and not re.match(r"^\w+\s*=", line)]
        has_rules = any(line.startswith("-") for line in rules)
        result = "符合" if has_rules else "不符合"
        finding = "auditd.conf 已输出；audit.rules 未发现有效审计规则。" if not has_rules else "auditd.conf 与 audit.rules 均存在有效配置。"
    elif row == 22:
        service_count = re.search(r"(\d+)\s+loaded units listed", text)
        risky_tokens = sorted({token.lower() for token in re.findall(r"\b(nginx|rpcbind|avahi-daemon|cups|X|telnet|ftp|rsh|vnc)\b", text, flags=re.I)})
        ports = sorted(set(re.findall(r"[:*](\d{2,5})\b", text)))
        result = "不符合" if risky_tokens or any(port in {"80", "111", "6000"} for port in ports) else "符合"
        finding = f"运行服务数量：{service_count.group(1) if service_count else '未识别'}；监听端口包括：{', '.join(ports[:10]) or '未识别'}。" + (f" 发现需确认的服务/端口：{', '.join(risky_tokens)}。" if risky_tokens else "")
    elif row == 23:
        has_access_rule = bool(re.search(r"^(?!PasswordAuthentication\b).*(AllowUsers|DenyUsers|AllowGroups|DenyGroups|ListenAddress|ALL:)", text, flags=re.I | re.M))
        result = "符合" if has_access_rule else "不符合"
        finding = "已配置终端登录来源或用户访问限制。" if has_access_rule else "hosts.allow/hosts.deny 未见有效限制规则；sshd_config 未见 Allow/Deny 访问控制。"
    elif row == 24:
        has_tmout = bool(re.search(r"\bTMOUT\s*=", text))
        result = "符合" if has_tmout else "不符合"
        finding = "已发现 TMOUT 超时锁定设置。" if has_tmout else "未发现 TMOUT 超时锁定设置。"
    else:
        lines = clean_lines(text)
        result = "符合" if lines else "未检查"
        finding = "；".join(lines[:4]) or "未获取到命令输出。"

    return finding, normalize_result(result, options)


def build_plan(
    workbook: Path,
    commands_json: Path,
    task_label: str,
    manifest_json: Path | None,
    finding_template: str | None,
    result_text: str | None,
    include_screenshots: bool,
) -> dict[str, Any]:
    wb = openpyxl.load_workbook(workbook, data_only=False)
    sheet = wb.active
    header_row, columns = find_header(sheet)
    result_options = result_options_for_sheet(sheet, columns.get("result"))
    commands = load_json_array(commands_json, "commands-json")
    manifest = manifest_by_id(manifest_json)

    items: list[dict[str, Any]] = []
    unmapped: list[str] = []
    mapped_rows: set[int] = set()

    for index, command in enumerate(commands, start=1):
        row = command_row(command)
        command_id = safe_file_part(str(command.get("id") or f"cmd{index:02d}"))
        if row is None:
            unmapped.append(command_id)
            continue
        if row <= header_row or row > sheet.max_row:
            raise SystemExit(f"Command {command_id!r} maps to row {row}, outside workbook data rows")
        mapped_rows.add(row)

        source = row_text(sheet, row, columns)
        name = norm(command.get("name")) or command_id
        command_text = norm(command.get("command"))
        artifact = command_artifact(command_id, command, manifest)
        output_text = command_output_text(command_id, command, manifest, manifest_json)
        summary_finding, inferred_result = summarize_linux_output(row, output_text, result_options)
        if command.get("finding"):
            finding = str(command["finding"])
        elif finding_template:
            finding = finding_template.format(
                id=command_id,
                name=name,
                command=command_text,
                row=row,
                artifact=artifact,
            )
        else:
            finding = summary_finding
        explicit_result = command.get("result") or result_text
        if explicit_result:
            result = normalize_result(
                str(explicit_result),
                result_options,
                strict=True,
                label=f"result for command {command_id!r}",
            )
        else:
            result = normalize_result(str(inferred_result), result_options)
        evidence = [evidence_name(index, command, manifest)] if include_screenshots else []

        items.append(
            {
                "row": row,
                "lane": "ssh",
                "confidence": "high",
                "check_id": command_id,
                "tool": "ssh",
                "gui_action": "ssh_command_screenshot" if include_screenshots else "ssh_command",
                "graphical_required": False,
                "screenshots_requested": include_screenshots,
                "evidence": evidence,
                "keywords": [],
                "source": source,
                "finding": finding,
                "result": result,
                "notes": "Linux/Unix SSH command evidence mapped to workbook output.",
                "ssh_command": command_text,
                "artifact": artifact,
            }
        )

    last_category = ""
    for row in range(header_row + 1, sheet.max_row + 1):
        if row in mapped_rows:
            continue
        source = row_text(sheet, row, columns)
        if source.get("category"):
            last_category = source["category"]
        elif last_category:
            source["category"] = last_category
        if not is_admin_interview(source):
            continue
        items.append(
            {
                "row": row,
                "lane": "skipped",
                "confidence": "high",
                "check_id": "admin_interview",
                "tool": None,
                "gui_action": "skip_admin_interview",
                "graphical_required": False,
                "screenshots_requested": include_screenshots,
                "evidence": [],
                "keywords": [],
                "source": source,
                "finding": ADMIN_INTERVIEW_FINDING,
                "result": ADMIN_INTERVIEW_RESULT,
                "notes": "Administrator interview requires offline IT administrator communication; do not operate.",
            }
        )

    if unmapped:
        raise SystemExit(
            "Commands could not be mapped to workbook rows. Add an explicit row/workbookRow "
            f"field or use ids like row05: {', '.join(unmapped)}"
        )

    return {
        "source_workbook": str(workbook),
        "task_label": task_label,
        "sheet": sheet.title,
        "header_row": header_row,
        "columns": columns,
        "result_options": result_options,
        "items": items,
        "summary": {
            "total_items": len(items),
            "ssh": sum(1 for item in items if item["lane"] == "ssh"),
            "skipped": sum(1 for item in items if item["lane"] == "skipped"),
            "screenshots": include_screenshots,
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-workbook", required=True, type=Path)
    parser.add_argument("--commands-json", required=True, type=Path)
    parser.add_argument("--task-label", required=True)
    parser.add_argument("--out", required=True, type=Path)
    parser.add_argument("--manifest-json", type=Path)
    parser.add_argument("--screenshots", action="store_true", help="Map screenshot filenames into the workbook plan.")
    parser.add_argument("--finding-template")
    parser.add_argument("--result-text", default=DEFAULT_RESULT_TEXT)
    args = parser.parse_args()

    if args.out.parent.name.lower() != "tmp":
        raise SystemExit("SSH workbook plan JSON must be written under an evidence tmp directory")

    finding_template = args.finding_template
    plan = build_plan(
        workbook=args.source_workbook,
        commands_json=args.commands_json,
        task_label=args.task_label,
        manifest_json=args.manifest_json,
        finding_template=finding_template,
        result_text=args.result_text,
        include_screenshots=args.screenshots,
    )
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(plan, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(plan["summary"], ensure_ascii=False))


if __name__ == "__main__":
    main()
