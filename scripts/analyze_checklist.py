#!/usr/bin/env python3
"""Analyze a Windows checklist workbook and produce an execution plan.

This script intentionally handles only planning. The agent remains responsible
for reviewing low-confidence/adaptive rows before execution.
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl is required to analyze .xlsx files") from exc


COLUMN_ALIASES = {
    "category": ["分类", "类别", "控制域"],
    "item": ["测评项", "检查项", "测评内容", "检查内容", "安全要求"],
    "expected": ["预期结果", "预期", "标准", "要求", "合规要求"],
    "operation": ["评估操作示例", "操作示例", "检查方法", "测评方法", "核查方法"],
    "finding": ["检查情况", "检查结果说明", "检查记录", "实际情况"],
    "result": ["结果", "符合性", "判定", "检查结果"],
    "remediation": ["整改建议", "修复建议", "加固建议"],
}

TOOL_HINTS = [
    ("secpol.msc", ["secpol.msc", "本地安全策略", "密码策略", "账户锁定", "审核策略", "安全选项"]),
    ("gpedit.msc", ["gpedit.msc", "组策略", "管理模板", "远程桌面", "终端服务"]),
    ("lusrmgr.msc", ["lusrmgr.msc", "用户", "组", "guest", "administrator", "来宾"]),
    ("services.msc", ["services.msc", "服务", "启动类型", "登录身份", "log on"]),
    ("eventvwr.msc", ["eventvwr", "eventvwr.msc", "事件查看器", "系统日志", "日志"]),
    ("fsmgmt.msc", ["fsmgmt.msc", "共享", "默认共享"]),
    ("regedit", ["regedit", "注册表", "hkey", "restrictanonymous"]),
    ("control", ["control", "控制面板", "已安装更新", "添加删除程序", "程序和功能"]),
]

KEYWORD_STOPWORDS = {
    "应", "应当", "检查", "核查", "测评", "要求", "是否", "进行", "相关", "系统", "用户", "管理", "配置",
    "the", "and", "for", "with", "shall", "should", "check", "verify",
}

ADMIN_INTERVIEW_FINDING = "涉及访谈管理员，未进行操作"
ADMIN_INTERVIEW_RESULT = "未检查"
ADMIN_INTERVIEW_TERMS = ("访谈", "询问", "问询", "沟通", "线下确认")
ADMIN_ROLE_TERMS = ("管理员", "管理人员", "系统管理员", "安全管理员", "it管理员", "it 管理员")


def norm(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\u3000", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def lower_text(value: str) -> str:
    return norm(value).lower()


def find_header(sheet: Any) -> tuple[int, dict[str, int]]:
    best: tuple[int, dict[str, int]] | None = None
    best_score = 0
    for row in range(1, min(sheet.max_row, 12) + 1):
        values = [norm(sheet.cell(row, col).value) for col in range(1, sheet.max_column + 1)]
        mapping: dict[str, int] = {}
        score = 0
        for key, aliases in COLUMN_ALIASES.items():
            used_columns = set(mapping.values())
            matched_col = None
            for col, value in enumerate(values, start=1):
                if col in used_columns or not value:
                    continue
                if any(alias == value for alias in aliases):
                    matched_col = col
                    break
            if matched_col is None:
                for col, value in enumerate(values, start=1):
                    if col in used_columns or not value:
                        continue
                    if any(alias in value for alias in aliases):
                        matched_col = col
                        break
            if matched_col is not None:
                mapping[key] = matched_col
                score += 1
        if score > best_score:
            best = (row, mapping)
            best_score = score
    if not best or best_score < 3:
        raise SystemExit("Could not identify checklist header row and columns")
    return best


def row_text(sheet: Any, row: int, columns: dict[str, int]) -> dict[str, str]:
    return {
        key: norm(sheet.cell(row, col).value)
        for key, col in columns.items()
        if col <= sheet.max_column
    }


def score_pattern(text: str, pattern: dict[str, Any]) -> int:
    match = pattern.get("match", {})
    score = 0
    for token in match.get("any", []):
        if lower_text(token) in text:
            score += 1
    all_tokens = match.get("all", [])
    if all_tokens and all(lower_text(token) in text for token in all_tokens):
        score += len(all_tokens) + 2
    for regex in match.get("regex", []):
        if re.search(regex, text, flags=re.I):
            score += 2
    return score


def infer_tool(text: str) -> str | None:
    for tool, hints in TOOL_HINTS:
        if any(lower_text(hint) in text for hint in hints):
            return tool
    return None


def is_admin_interview(data: dict[str, str]) -> bool:
    combined = lower_text(" ".join(data.get(k, "") for k in ("category", "item", "expected", "operation", "remediation")))
    return any(term in combined for term in ADMIN_INTERVIEW_TERMS) and any(term in combined for term in ADMIN_ROLE_TERMS)


def extract_keywords(data: dict[str, str], pattern: dict[str, Any] | None) -> list[str]:
    """Extract UI navigation/validation keywords from the checklist row text."""
    combined = " ".join(data.get(k, "") for k in ("item", "expected", "operation", "remediation", "category"))
    combined_lower = lower_text(combined)
    keywords: list[str] = []

    if pattern:
        for token in pattern.get("match", {}).get("any", []):
            token = norm(token)
            if token and "?" not in token and lower_text(token) in combined_lower:
                keywords.append(token)

    for token in re.findall(r"[\u4e00-\u9fffA-Za-z0-9_$\\.-]{2,}", combined):
        token = norm(token)
        token_lower = lower_text(token)
        if token_lower in KEYWORD_STOPWORDS:
            continue
        if len(token) > 18 and not re.search(r"[A-Za-z0-9_$\\.-]", token):
            continue
        keywords.append(token)

    deduped: list[str] = []
    seen: set[str] = set()
    for token in keywords:
        key = lower_text(token)
        if key in seen:
            continue
        seen.add(key)
        deduped.append(token)
    return deduped[:20]


def evidence_names(row: int, pattern: dict[str, Any] | None, tool: str | None) -> list[str]:
    if pattern:
        slugs = [pattern["evidence_slug"]]
        slugs.extend(pattern.get("extra_evidence_slugs", []))
        return [f"row{row:02d}_{slug}.png" for slug in slugs]
    tool_slug = (tool or "adaptive").replace(".msc", "").replace(";", "_").replace(" ", "_")
    return [f"row{row:02d}_{tool_slug}_adaptive.png"]


def plan_rows(workbook: Path, patterns_path: Path, task_label: str, include_screenshots: bool = False) -> dict[str, Any]:
    patterns = json.loads(patterns_path.read_text(encoding="utf-8"))
    wb = openpyxl.load_workbook(workbook, data_only=False)
    sheet = wb.active
    header_row, columns = find_header(sheet)

    items: list[dict[str, Any]] = []
    last_category = ""
    for row in range(header_row + 1, sheet.max_row + 1):
        data = row_text(sheet, row, columns)
        if data.get("category"):
            last_category = data["category"]
        data.setdefault("category", last_category)
        combined = " ".join(data.get(k, "") for k in ("category", "item", "expected", "operation", "remediation"))
        combined_lower = lower_text(combined)
        if not combined_lower:
            continue
        if len(combined_lower) < 6:
            continue

        admin_interview = is_admin_interview(data)
        if admin_interview:
            lane = "skipped"
            confidence = "high"
            pattern = None
            tool = None
            action = "skip_admin_interview"
            evidence = []
            keywords = []
            notes = "Administrator interview requires offline IT administrator communication; do not operate."
            check_id = "admin_interview"
            graphical_required = False
            finding = ADMIN_INTERVIEW_FINDING
            result = ADMIN_INTERVIEW_RESULT
        else:
            scored = sorted(((score_pattern(combined_lower, p), p) for p in patterns), key=lambda pair: pair[0], reverse=True)
            best_score, best_pattern = scored[0]
            tool = infer_tool(combined_lower)

            if best_score >= 2:
                lane = "known"
                confidence = "high" if best_score >= 3 else "medium"
                pattern = best_pattern
                tool = pattern.get("tool") or tool
                action = pattern.get("gui_action")
            elif tool:
                lane = "inferred"
                confidence = "medium"
                pattern = None
                action = "adaptive_gui"
            else:
                lane = "adaptive"
                confidence = "low"
                pattern = None
                action = "adaptive_gui"
            evidence = evidence_names(row, pattern, tool) if include_screenshots else []
            keywords = extract_keywords(data, pattern)
            notes = pattern.get("notes") if pattern else "Use adaptive GUI rules; do not skip by default."
            check_id = pattern["id"] if pattern else None
            graphical_required = True if not pattern else bool(pattern.get("graphical_required", True))
            finding = None
            result = None

        item = {
            "row": row,
            "lane": lane,
            "confidence": confidence,
            "check_id": check_id,
            "tool": tool,
            "gui_action": action,
            "graphical_required": graphical_required,
            "screenshots_requested": include_screenshots,
            "evidence": evidence,
            "keywords": keywords,
            "source": data,
            "notes": notes,
        }
        if finding is not None:
            item["finding"] = finding
        if result is not None:
            item["result"] = result

        items.append(item)

    return {
        "source_workbook": str(workbook),
        "task_label": task_label,
        "sheet": sheet.title,
        "header_row": header_row,
        "columns": columns,
        "items": items,
        "summary": {
            "total_items": len(items),
            "known": sum(1 for item in items if item["lane"] == "known"),
            "inferred": sum(1 for item in items if item["lane"] == "inferred"),
            "adaptive": sum(1 for item in items if item["lane"] == "adaptive"),
            "skipped": sum(1 for item in items if item["lane"] == "skipped"),
            "screenshots": include_screenshots,
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--patterns", required=True, type=Path)
    parser.add_argument("--task-label", required=True)
    parser.add_argument("--out", required=True, type=Path)
    parser.add_argument("--screenshots", action="store_true", help="Include final screenshot filenames in the plan.")
    args = parser.parse_args()

    if args.out.parent.name.lower() != "tmp":
        raise SystemExit("Execution plan JSON must be written under an evidence tmp directory, for example <evidence>/tmp/plan.json")

    plan = plan_rows(args.workbook, args.patterns, args.task_label, include_screenshots=args.screenshots)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(plan, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(plan["summary"], ensure_ascii=False))


if __name__ == "__main__":
    main()
