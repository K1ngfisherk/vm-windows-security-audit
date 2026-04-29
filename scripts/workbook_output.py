#!/usr/bin/env python3
"""Write checklist results into a copied workbook.

Modes:
- text: write finding/status and plain evidence filenames.
- embed-images: write finding/status and insert images in the finding column.

The input plan may be enriched after execution with `finding`, `status`, and
`evidence` fields per item. This script keeps the original workbook untouched.
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.drawing.image import Image
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl and Pillow are required for workbook output") from exc


ADMIN_INTERVIEW_FINDING = "涉及访谈管理员，未进行操作"
ADMIN_INTERVIEW_RESULT = "未检查"
STATUS_RESULT_ALIASES = {
    "已检查",
    "通过",
    "合规",
    "符合要求",
    "不通过",
    "不合规",
    "不符合要求",
    "na",
    "n/a",
    "不涉及",
    "未执行",
    "未操作",
}


def col_letter(index: int) -> str:
    return openpyxl.utils.get_column_letter(index)


def write_text(cell: Any, text: str) -> None:
    cell.value = text.strip() if text else ""


def strip_evidence_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n")
    text = re.sub(r"\n?截图[:：][\s\S]*$", "", text)
    text = re.sub(r"\n?证据[:：][\s\S]*$", "", text)
    text = re.sub(r"\n?row\d{2}_[^\s；;，,]+\.png", "", text, flags=re.I)
    return text.strip()


def normalize_result(value: str, options: list[str]) -> str:
    text = value.strip()
    if not options or not text:
        return text
    if text in options:
        return text
    mapping = {
        "已检查": "符合",
        "通过": "符合",
        "合规": "符合",
        "符合要求": "符合",
        "不通过": "不符合",
        "不合规": "不符合",
        "不符合要求": "不符合",
        "na": "不适用",
        "n/a": "不适用",
        "不涉及": "不适用",
        "未执行": "未检查",
        "未操作": "未检查",
    }
    mapped = mapping.get(text.lower(), mapping.get(text))
    if mapped and mapped in options:
        return mapped
    for option in options:
        if option in text or text in option:
            return option
    return "符合" if "符合" in options else options[0]


def report_result(item: dict[str, Any], options: list[str] | None = None) -> str:
    options = options or []
    if is_admin_interview_skip(item):
        return normalize_result(ADMIN_INTERVIEW_RESULT, options)
    for key in ("result", "compliance", "judgement", "judgment"):
        if item.get(key):
            return normalize_result(str(item[key]), options)
    status = str(item.get("status") or "")
    if status and (status in options or status in STATUS_RESULT_ALIASES):
        return normalize_result(status, options)
    return ""


def is_admin_interview_skip(item: dict[str, Any]) -> bool:
    return (
        item.get("check_id") == "admin_interview"
        or item.get("gui_action") == "skip_admin_interview"
        or item.get("skip_reason") == "administrator_interview"
    )


def merge_runner_result(plan: dict[str, Any], runner_result: dict[str, Any]) -> dict[str, Any]:
    results_by_row = {
        int(result["row"]): result
        for result in runner_result.get("results", [])
        if result.get("row") not in (None, "")
    }
    for item in plan.get("items", []):
        row = int(item["row"])
        result = results_by_row.get(row)
        if not result:
            continue
        for key, value in result.items():
            if key == "row":
                continue
            item[key] = value
    return plan


def validate_written_workbook(ws: Any, plan: dict[str, Any], finding_col: int, result_col: int, result_options: list[str]) -> None:
    errors: list[str] = []
    for item in plan.get("items", []):
        row = int(item["row"])
        expected_finding = item.get("finding") or item.get("observed") or ""
        if is_admin_interview_skip(item):
            expected_finding = ADMIN_INTERVIEW_FINDING
        expected_result = report_result(item, result_options)
        actual_finding = strip_evidence_text(ws.cell(row, finding_col).value)
        actual_result = str(ws.cell(row, result_col).value or "").strip()
        if expected_finding and expected_finding.strip() not in actual_finding:
            errors.append(f"row {row} finding mismatch")
        if expected_result and actual_result != expected_result:
            errors.append(f"row {row} result mismatch: expected {expected_result!r}, got {actual_result!r}")
    if errors:
        raise SystemExit("Workbook validation failed after write: " + "; ".join(errors[:8]))


def cleanup_tmp_after_report(plan_path: Path, evidence_dir: Path) -> None:
    tmp_dir = plan_path.parent.resolve()
    evidence_root = evidence_dir.resolve()
    if tmp_dir.name.lower() != "tmp":
        raise SystemExit(f"Refusing to clean non-tmp plan parent: {tmp_dir}")
    try:
        tmp_dir.relative_to(evidence_root)
    except ValueError as exc:
        raise SystemExit(f"Refusing to clean tmp outside evidence dir: {tmp_dir}") from exc
    shutil.rmtree(tmp_dir)


def fit_image(img: Image, max_width: int, max_height: int) -> Image:
    scale = min(max_width / img.width, max_height / img.height, 1.0)
    img.width = int(img.width * scale)
    img.height = int(img.height * scale)
    return img


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-workbook", required=True, type=Path)
    parser.add_argument("--plan", required=True, type=Path)
    parser.add_argument("--evidence-dir", required=True, type=Path)
    parser.add_argument("--output-workbook", required=True, type=Path)
    parser.add_argument("--mode", choices=["text", "embed-images"], default="text")
    parser.add_argument("--runner-result", type=Path, help="Optional runner_result.json to merge into the plan before writing.")
    parser.add_argument("--cleanup-tmp", action="store_true", help="Remove evidence/tmp after workbook validation succeeds.")
    args = parser.parse_args()

    if args.source_workbook.resolve() != args.output_workbook.resolve():
        args.output_workbook.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(args.source_workbook, args.output_workbook)

    plan = json.loads(args.plan.read_text(encoding="utf-8"))
    if args.runner_result:
        runner_result = json.loads(args.runner_result.read_text(encoding="utf-8"))
        plan = merge_runner_result(plan, runner_result)
    wb = openpyxl.load_workbook(args.output_workbook)
    ws = wb[plan["sheet"]]
    finding_col = int(plan["columns"].get("finding") or 5)
    result_col = int(plan["columns"].get("result") or 6)
    result_options = [str(value) for value in plan.get("result_options") or [] if str(value).strip()]

    for item in plan["items"]:
        row = int(item["row"])
        finding = item.get("finding") or item.get("observed") or strip_evidence_text(ws.cell(row, finding_col).value)
        if is_admin_interview_skip(item):
            finding = ADMIN_INTERVIEW_FINDING
        status = report_result(item, result_options)
        evidence = [] if is_admin_interview_skip(item) else item.get("evidence") or []

        if args.mode == "text":
            evidence_text = "\n".join(f"截图：{name}" for name in evidence)
            write_text(ws.cell(row, finding_col), "\n".join(part for part in [finding, evidence_text] if part))
            if status:
                write_text(ws.cell(row, result_col), status)
            continue

        write_text(ws.cell(row, finding_col), finding)
        if status:
            write_text(ws.cell(row, result_col), status)

        left_anchor = f"{col_letter(finding_col)}{row}"
        max_image_height = 245 if len(evidence) > 1 else 320

        for index, name in enumerate(evidence):
            image_path = args.evidence_dir / name
            if not image_path.exists():
                continue
            img = fit_image(Image(str(image_path)), 300 if len(evidence) > 1 else 520, max_image_height)
            img.anchor = left_anchor
            ws.add_image(img)
            # openpyxl does not provide simple cell-internal offsets here.
            # For high-fidelity placement, prefer Excel COM or artifact-tool.
            if index > 0:
                break

    wb.save(args.output_workbook)
    validate_written_workbook(ws, plan, finding_col, result_col, result_options)
    if args.cleanup_tmp:
        cleanup_tmp_after_report(args.plan, args.evidence_dir)
    print(json.dumps({"output_workbook": str(args.output_workbook), "mode": args.mode, "tmp_removed": bool(args.cleanup_tmp)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
